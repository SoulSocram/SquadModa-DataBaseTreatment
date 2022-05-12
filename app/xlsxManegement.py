from openpyxl import Workbook, load_workbook

class xlsx:

    def readCell(file, cell):

        return file[cell].value

    def writeCell(file, cell, var):
        print(cell)
        print(file[f'{cell}'].value)
        file[cell].value = var

    def openFile(path):

        wb = load_workbook(path)
        return wb

    def openWorkbook(file, workbook):

        p1 = file[workbook]
        return p1

    def save(file, path):

        file.save(path)

    def close(file):

        file.close()